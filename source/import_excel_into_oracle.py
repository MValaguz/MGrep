# -*- coding: utf-8 -*-

"""
 Creato da.....: Marco Valaguzza
 Piattaforma...: Python3.6
 Data..........: 24/01/2020
 Descrizione...: Scopo dello script è prendere un foglio di excel e caricarlo in un DB Oracle.
 Note..........: Al momento la procedura non permette di personalizzare il tipo e il numero di colonne da importare.
                 Si dà per scontato che alla prima riga siano presenti i nomi delle colonne.
                 Il programma per ogni colonna ricerca la lunghezza massima del campo.
                 Ricerca per ogni colonna il tipo predominante. Nel caso in cui i valori non predominanti risulteranno
                 incopatibili con i valori dominanti, ci potrebbero essere degli errori.
"""

#Librerie di data base
import  cx_Oracle
#Librerie di sistema
import os
import sys
#Librerie grafiche
from PyQt5 import QtCore, QtGui, QtWidgets
#Libreria per la lettura e scrittura file di excel
from    openpyxl import load_workbook
#Liberia regular expression
import  re
#Moduli di progetto
from utilita import message_error, message_info, message_question_yes_no

def slugify(text, lower=1):
    """
        questa funzione usando le regular expression, normalizza il nome di una colonna 
        togliendo caratteri indesiderati
    """
    if lower == 1:
        text = text.strip().lower()
    text = re.sub(r'[^\w _-]+', '', text)
    text = re.sub(r'[- ]+', '_', text)
    return text

class import_excel_into_oracle(QtWidgets.QWidget):
    """
        Importa un foglio di excel dentro una tabella di Oracle
        Va indicato attraverso l'instanziazione della classe:
            p_debug          = Se True --> esegue le print del lavoro svolto
            p_user_db        = Nome utente del DB Oracle
            p_password_db    = Password utente del DB Oracle
            p_dsn_db         = Indirizzo IP del DB Oracle o dsn
            p_table_name     = Nome della tabella di destinazione Oracle
            p_excel_file     = Nome del file di excel (compresa di pathname)        
            p_modalita_test  = Modalita test
    """                 
    def __init__(self,
                 p_debug,
                 p_user_db,
                 p_password_db,
                 p_dsn_db,
                 p_table_name,
                 p_excel_file,
                 p_modalita_test):
                
        # rendo la mia classe una superclasse
        super(import_excel_into_oracle, self).__init__()                
        
                # creazione della wait window
        self.v_progress_step = 0
        self.progress = QtWidgets.QProgressDialog(self)        
        self.progress.setMinimumDuration(0)
        self.progress.setWindowModality(QtCore.Qt.WindowModal)
        self.progress.setWindowTitle("Copy...")                
        
        # icona di riferimento
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(":/icons/icons/MGrep.ico"), QtGui.QIcon.Normal, QtGui.QIcon.Off)        
        self.progress.setWindowIcon(icon)
        
        # imposto valore minimo e massimo a 0 in modo venga considerata una progress a tempo indefinito
        # Attenzione! dentro nel ciclo deve essere usata la funzione setvalue altrimenti non visualizza e non avanza nulla!
        self.progress.setMinimum(0)
        self.progress.setMaximum(100) 
        # creo un campo label che viene impostato con 100 caratteri in modo venga data una dimensione di base standard
        self.progress_label = QtWidgets.QLabel()            
        self.progress_label.setText('.'*100)
        # collego la label già presente nell'oggetto progress bar con la mia label 
        self.progress.setLabel(self.progress_label)                           
        # creo una scritta iniziale...
        self.avanza_progress("Analizing excel file...")
        self.avanza_progress("Analizing excel file...")
        
        #Apro il file di excel
        try:
            wb = load_workbook(filename = p_excel_file)
        except:
            message_error('Format file invalid. Only xlsx file format!')
            #esco
            return None
        #Leggo tutti i nomi dei fogli in esso contenuti
        v_nomi_fogli = wb.sheetnames
        #Se il file contiene più fogli, avverto che verrà preso solo il primo
        if len(v_nomi_fogli) > 1:
            message_info('This file contains more than one sheet. It will be taken the first ' + v_nomi_fogli[0] + '!')            
        
        #Mi posiziono sul primo foglio 
        v_foglio = wb[v_nomi_fogli[0]]
        
        #Estraggo la struttura del foglio
        self.v_numero_totale_righe = 0
        self.v_debug = p_debug
        v_definizione_colonne = self.struttura_foglio(v_foglio) 
        
        self.avanza_progress( 'Total records number to copy...' + str(self.v_numero_totale_righe) )
        
        #Collegamento a Oracle                
        try:
            self.v_oracle_db = cx_Oracle.connect(user=p_user_db, password=p_password_db, dsn=p_dsn_db, encoding = "UTF-8", nencoding = "UTF-8")
        except:
            message_error('Connecting problems to Oracle DB!')
            #esco
            return None
        
        self.v_oracle_cursor = self.v_oracle_db.cursor()    
        
        #Creo la tabella                                 
        if self.creo_tabella(v_definizione_colonne, p_table_name) == 'ok':
            if self.importa_foglio(v_foglio, v_definizione_colonne, p_table_name) == 'ok':
                #Messaggio finale
                message_info('Action completed with ' + str(self.v_numero_totale_righe) + ' records imported!')        
        
        self.v_oracle_db.close()                
        return None        
    
    def struttura_foglio(self, p_foglio):        
        #Leggo il foglio per colonne, in questo modo ottengo come risultante la struttura della tabella da creare
        #tenendo conto della larghezza massima di ogni colonna. La lista "v_definizione_colonne" conterrà per ogni riga,
        #4 campi che saranno, il nome, il tipo, la lunghezza e eventuali decimali se tipo è numerico
        v_definizione_colonne = []  
        self.v_numero_totale_righe = 0      
        v_numero_colonna = 0
        for col in p_foglio.iter_cols():    
            v_1a_riga = True
            v_numero_colonna += 1
            v_nome_colonna = ''
            #questa lista serve per determinare il tipo di campo "prevalente" che è presente nella colonna. 
            #non verrà preso il primo tipo campo che capita ma quello che risulta avere il risultato più alto tra tutti
            #i tipi campo trovati
            v_tipo_colonna = 'VARCHAR2'                
            v_varchar2 = 0
            v_number = 0
            v_date = 0
            v_other = 0
            v_larghezza_colonna = 1
            v_larghezza_decimali = 0            
            v_numero_righe_per_colonna = 0
            for cell in col:
                #la prima riga contiene per standard il nome della colonna
                if v_1a_riga:
                    v_1a_riga = False
                    if cell.value is not None:
                        #Estraggo e normalizzo il nome della colonna 
                        v_nome_colonna = slugify(cell.value)
                        #Se nome della colonna supera i 30 caratteri --> lo tronco a 30
                        if len(v_nome_colonna) > 30:
                            v_nome_colonna = v_nome_colonna[0:30]
                    else:
                        v_nome_colonna = 'COL_' + str(v_numero_colonna)
                                
                #per le altre righe devo calcolare il tipo di dato e la larghezza massima 
                #il tipo di dato dipende dal primo valore che incontra....dovrà essere eventualmente perfezionata questa cosa 
                else:
                    #considero solo celle che contengono un valore (escludo anche quelle con una specie di spazio all'interno)
                    if cell.value is not None and str(cell.value) != chr(160) and cell.data_type != 'f':                        
                        #for i in range(1,255):
                        #    if str(cell.value) == chr(i):
                        #        print('-' + str(cell.value) + '-' + str(i))                        
                        #la cella è di tipo testo (VARCHAR2)                        
                        if cell.data_type == 's':                                                        
                            v_varchar2 += 1
                            if len(cell.value) > v_larghezza_colonna:
                                v_larghezza_colonna = len(cell.value)
                        #la cella è numerica (NUMBER)
                        elif cell.data_type == 'n':                            
                            v_number += 1
                            v_stringa = str(cell.value)                            
                            #normalizzo virgola con punto
                            v_stringa = v_stringa.replace(',','.')
                            #è presente la parte decimale (divido la parte intera dalla decimale)
                            if v_stringa.find('.') > 0:  
                                if len(v_stringa.split('.')[0]) > v_larghezza_colonna:
                                    v_larghezza_colonna = len(v_stringa.split('.')[0])
                                if len(v_stringa.split('.')[1]) > v_larghezza_decimali:
                                    v_larghezza_decimali = len(v_stringa.split('.')[1])
                             #è presente solo la parte intera
                            elif len(v_stringa) > v_larghezza_colonna:
                                v_larghezza_colonna = len(v_stringa)                            
                        #la cella è una data (DATE)
                        elif cell.data_type == 'd':
                            v_date += 1
                            v_larghezza_colonna = 20
                        #la cella non è riconoscibile (OTHER)
                        else:
                            v_other += 1
                #aggiorno numero righe trovate nella colonna
                v_numero_righe_per_colonna += 1
                                
                    
            #controllo qual'è il risultato del tipo colonna prevalente (di base varchar2)                        
            #una colonna viene definita numerica o di data solo se nella colonna non ci sono caratteri
            if v_varchar2 == 0:
                if v_number > 0:
                    v_tipo_colonna = 'NUMBER'
                elif v_date > 0:
                    v_tipo_colonna = 'DATE'                            
            
            #salvo la lista dei dati di definizione della colonna
            if self.v_debug:
                print("Definizione delle colonne Nome: " + v_nome_colonna + " tipo: " + v_tipo_colonna + " larghezza: " +  str(v_larghezza_colonna) + " decimali: " + str(v_larghezza_decimali))                

            v_definizione_colonne.append((v_nome_colonna, v_tipo_colonna, v_larghezza_colonna, v_larghezza_decimali))
            
            #aggiorno il numero delle righe contenuto del foglio (è il valore massimo delle righe trovate su tutte le colonne)
            if v_numero_righe_per_colonna > self.v_numero_totale_righe:
                self.v_numero_totale_righe = v_numero_righe_per_colonna
                                    
        return v_definizione_colonne
    
    def creo_tabella(self, 
                     p_definizione_colonne, 
                     p_table_name):
         
        #se siamo in modalità test --> cancello la tabella
        if self.v_debug:
            try:
                print("Cancellazione della tabella " + p_table_name)                
                v_query = 'DROP TABLE ' + p_table_name
                self.v_oracle_cursor.execute(v_query)
            except:
                pass
                        
        v_query = 'CREATE TABLE ' + p_table_name + '('
        v_1a_volta = True
        #definizione_colonne risulta così strutturata
        #1° campo = Nome
        #2° campo = Tipo (varchar2, number)
        #3° campo = Larghezza colonna (per numeri solo la parte intera)
        #4° campo = Larghezza decimali (solo per numeri con decimali)
        for valori in p_definizione_colonne:
            if v_1a_volta:
                v_1a_volta=False
            else:
                v_query += ','
            v_query += valori[0] + ' ' + valori[1]
            if valori[1] == 'VARCHAR2':
                #la colonna risulta varchar2 ma sono stati trovati all'interno dei numeri con decimali....per evitare
                #problemi aumento la grandezza della colonna aggiungendo anche i decimali + altri due caratteri per 
                #separatore decimali e eventuale segno meno
                if valori[3] > 0:
                    v_query += '(' + str(valori[2]+valori[3]+2) + ')'
                #altrimenti è puro char
                else:
                    v_query += '(' + str(valori[2]) + ')'
            elif valori[1] == 'NUMBER':
                v_query += '(' + str(valori[2]+valori[3]) + ',' + str(valori[3]) + ')'            
        v_query += ')'
        
        if self.v_debug:
            print("Creazione della tabella " + v_query)                
                    
        #Invio del comando di creazione tabella
        try:    
            self.v_oracle_cursor.execute(v_query)
        except:
            message_error("Problem during create Oracle table! The table " + p_table_name + " already exists?")            
            #esco
            return 'ko'

        return 'ok'
    
    def importa_foglio(self,                        
                       p_foglio, 
                       p_definizione_colonne,
                       p_table_name):
        #Calcolo 1% che rappresenta lo spostamento della progress bar
        v_rif_percent = 0
        if self.v_numero_totale_righe > 100:
            v_rif_percent = self.v_numero_totale_righe // 100

        #Leggo il foglio per righe        
        v_1a_riga = True
        v_progress = 0
        for row in p_foglio.iter_rows():    
            if not v_1a_riga:
                v_1a_volta = True
                v_insert = "INSERT INTO " + p_table_name + " VALUES("
                v_i = 0
                #Per ogni riga creo la relativa insert, tenendo conto del tipo di dato definito
                for cell in row:
                    if v_1a_volta:
                        v_1a_volta = False
                    else:
                        v_insert += ","
                    #se la cella è vuota o ti tipo formula --> null
                    if cell.value == None or cell.data_type == 'f':
                        v_insert += "null"
                    elif p_definizione_colonne[v_i][1] == 'VARCHAR2':
                        v_valore_stringa = str(cell.value)
                        #sostituisce il carattere unicode "rombo con il ?" con un asterisco
                        v_valore_stringa = v_valore_stringa.replace(u"\ufffd", "*")
                        #sostituisco il carattere apice con il doppio apice
                        v_valore_stringa = v_valore_stringa.replace("'","''")
                        v_insert += "'" + v_valore_stringa + "'"
                    elif p_definizione_colonne[v_i][1] == 'NUMBER':
                        #v_insert += "'" + str(cell.value).replace(',','.') + "'"
                        v_str_number = str(cell.value).strip()
                        if v_str_number == '':
                            v_insert += "null"
                        else:
                            v_insert += v_str_number
                    elif p_definizione_colonne[v_i][1] == 'DATE':
                        v_data = str(cell.value)
                        #normalizzo la data con le / a -
                        v_data = v_data.replace('/','-')
                        v_insert += "TO_DATE('" + v_data + "','RRRR-MM-DD HH24:MI:SS')"
                    v_i += 1
                v_insert += ")"        
                #eseguo la insert
                if self.v_debug:
                    print(v_insert)                                        
                                
                self.v_oracle_cursor.execute(v_insert)                
                #Emetto percentuale di avanzamento ma solo se righe maggiori di 100
                if self.v_numero_totale_righe > 100:
                    v_progress += 1
                    if v_progress % v_rif_percent == 0:                                        
                        self.avanza_progress( 'Total records to copy: ' + str(self.v_numero_totale_righe) )
                        
            else:
                v_1a_riga = False
                    
        self.v_oracle_db.commit()
        return 'ok'
    
    def avanza_progress(self, p_msg):
        """
           Visualizza prossimo avanzamento sulla progress bar
        """
        self.v_progress_step += 1
        if self.v_progress_step <= 100:
            self.progress.setValue(self.v_progress_step);                                        
            self.progress_label.setText(p_msg)         
        
# ----------------------------------------
# TEST APPLICAZIONE
# ----------------------------------------
if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)    
    import_excel_into_oracle(True,
                             "SMILE",
                             "SMILE",
                             "BACKUP_815",
                             "CANCELLAMI_91",                                         
                             "C:/MGrep/exportdb.xlsx",
                             True)      